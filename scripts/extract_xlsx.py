"""Extract CKA_Model_FINAL.xlsx into clean CSVs for the paper."""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "CKA_Model_FINAL.xlsx"
OUT = ROOT / "data"
OUT.mkdir(parents=True, exist_ok=True)

GROUPS = {
    "30 Models CONTROL BOOTSTRAP": "control",
    "30 Models CKA BOOTSTRAP": "cka",
    "30 Models COS BOOTSTRAP": "cos",
}

BENCHMARK_HEADERS = ["ARC-Challenge", "ARC-Easy", "GPQA", "GSM8K", "Human Eval", "MMLU"]
BENCHMARK_SLUG = {
    "ARC-Challenge": "arc_challenge",
    "ARC-Easy": "arc_easy",
    "GPQA": "gpqa",
    "GSM8K": "gsm8k",
    "Human Eval": "humaneval",
    "MMLU": "mmlu",
}


def cell(ws, row: int, col: int):
    """1-indexed cell value."""
    return ws.cell(row=row, column=col).value


def find_section_header(ws, name: str, search_col: int = 2):
    for r in range(1, ws.max_row + 1):
        v = cell(ws, r, search_col)
        if v and str(v).strip().rstrip(":") == name.rstrip(":"):
            return r
    return None


def extract_cka_matrix(ws, group_slug: str):
    """The matrix starts at row 11 with model_idx headers, then 30x30 numbers."""
    header_row = find_section_header(ws, "CKA Score Matrix")
    if header_row is None:
        return None


    cols_header_row = header_row + 1

    col_start = None
    for c in range(3, ws.max_column + 1):
        v = cell(ws, cols_header_row, c)
        try:
            float(v)
            col_start = c
            break
        except (TypeError, ValueError):
            continue
    if col_start is None:
        return None

    matrix = []
    for r in range(cols_header_row + 1, cols_header_row + 1 + 30):
        row_vals = []
        for c in range(col_start, col_start + 30):
            v = cell(ws, r, c)
            row_vals.append(float(v) if v is not None else float("nan"))
        matrix.append(row_vals)

    df = pd.DataFrame(matrix, columns=[f"m{i}" for i in range(30)],
                      index=[f"m{i}" for i in range(30)])
    out = OUT / f"cka_matrix_{group_slug}.csv"
    df.to_csv(out)
    print(f"  wrote {out.relative_to(ROOT)}  ({df.shape})")
    return df


def extract_benchmark_scaling(ws, group_slug: str):
    for bench in BENCHMARK_HEADERS:
        hdr_row = find_section_header(ws, bench)
        if hdr_row is None:
            continue

        col_hdr_row = hdr_row + 1

        headers = []
        cols = []
        for c in range(3, ws.max_column + 1):
            v = cell(ws, col_hdr_row, c)
            if v is None:
                if headers:
                    break
                continue
            headers.append(str(v))
            cols.append(c)
        rows_data = []
        for r in range(col_hdr_row + 1, col_hdr_row + 1 + 30):
            row = []
            for c in cols:
                v = cell(ws, r, c)
                row.append(v)
            if all(x is None for x in row):
                break
            rows_data.append(row)
        df = pd.DataFrame(rows_data, columns=headers)


        rename = {
            "Best Single Acc": "BestSingleAcc",
            "Added Model": "AddedModel",
            "Added Model Acc": "AddedModelAcc",
            "Added Acc": "AddedDeltaFromBase",
            "OracleAccAdded": "OracleDeltaFromBase",
        }
        df = df.rename(columns=rename)
        slug = BENCHMARK_SLUG[bench]
        out = OUT / f"scaling_{slug}_{group_slug}.csv"
        df.to_csv(out, index=False)
        print(f"  wrote {out.relative_to(ROOT)}  ({df.shape})")


def extract_router_results(ws, group_slug: str):
    hdr_row = find_section_header(ws, "Router Performance")
    if hdr_row is None:
        return
    col_hdr_row = hdr_row + 1

    headers = []
    cols = []
    for c in range(3, ws.max_column + 1):
        v = cell(ws, col_hdr_row, c)
        if v is None:
            if headers:
                break
            continue
        headers.append(str(v))
        cols.append(c)
    rows_data = []
    for r in range(col_hdr_row + 1, col_hdr_row + 25):
        row = []
        for c in cols:
            row.append(cell(ws, r, c))
        if all(x is None for x in row):
            break
        rows_data.append(row)
    df = pd.DataFrame(rows_data, columns=headers)
    out = OUT / f"router_{group_slug}.csv"
    df.to_csv(out, index=False)
    print(f"  wrote {out.relative_to(ROOT)}  ({df.shape})")


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    for sheet_name, slug in GROUPS.items():
        ws = wb[sheet_name]
        print(f"\n--- {sheet_name} -> {slug} ---")
        extract_cka_matrix(ws, slug)
        extract_benchmark_scaling(ws, slug)
        extract_router_results(ws, slug)


    ws = wb["Combined"]
    print("\n--- Combined ---")

    for label, slug in [("Bootstrap Router", "control_combined"),
                        ("CKA Router", "cka_combined"),
                        ("COS Performance", "cos_combined")]:
        hdr_row = find_section_header(ws, label)
        if hdr_row is None:
            continue
        col_hdr_row = hdr_row + 1
        headers = []
        cols = []
        for c in range(3, ws.max_column + 1):
            v = cell(ws, col_hdr_row, c)
            if v is None:
                if headers:
                    break
                continue
            headers.append(str(v))
            cols.append(c)
        rows_data = []
        for r in range(col_hdr_row + 1, col_hdr_row + 10):
            row = [cell(ws, r, c) for c in cols]
            if all(x is None for x in row):
                break
            rows_data.append(row)
        df = pd.DataFrame(rows_data, columns=headers)
        out = OUT / f"router_combined_{slug.replace('_combined','')}.csv"
        df.to_csv(out, index=False)
        print(f"  wrote {out.relative_to(ROOT)}  ({df.shape})")


if __name__ == "__main__":
    main()
